import os
import openpyxl
from openpyxl.utils import get_column_letter
from ir2tab_instr import extract_instructions_from_block
from ir2tab_function import extract_functions_from_ir
from ir2tab_block import extract_basic_blocks_from_function
from operator_formulas import build_dependency_formula
from classify_format import classify_format

def convert_to_spreadsheet_with_dependency(functions, output_xlsx):
    """
    Converts functions with basic blocks and instructions into a spreadsheet.
    Each instruction's result is represented in a cell with a formula showing dependencies
    using cell numbers.
    """
    # Ensure the output directory exists
    output_dir = os.path.dirname(output_xlsx)
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IR Dependencies"

    # Header row matching the desired format
    ws.append(["Instruction #", "Format", "Opcode", "Destination", "r1", "r2"])

    instruction_counter = 1  # Track instruction numbers

    for function_idx, function in enumerate(functions):
        basic_blocks = extract_basic_blocks_from_function(function)
        
        for block_idx, block in enumerate(basic_blocks):
            instructions = extract_instructions_from_block(block)
            
            for instr in instructions:
                opcode = instr['opcode']
                result_var = instr['result_var'] if instr['result_var'] else "NA"
                operands = instr['operands']

                # Classify format
                instr_format = classify_format(opcode)

                # Fill operands for r1 and r2
                r1 = operands[0] if len(operands) > 0 else "NA"
                r2 = operands[1] if len(operands) > 1 else "NA"

                # Add formula if it's a calculation (e.g., mul operation in R-format)
                if opcode == "mul":
                    r1_cell = f"E{ws.max_row + 1}"
                    r2_cell = f"F{ws.max_row + 1}"
                    formula = f"={r1_cell}*{r2_cell}"
                else:
                    formula = None

                # Write row data
                ws.append([instruction_counter, instr_format, opcode, result_var, r1, r2])

                # If there's a formula, insert it in the destination column
                if formula:
                    ws[f"D{ws.max_row}"] = formula

                instruction_counter += 1

    # Save the workbook
    wb.save(output_xlsx)
    print(f"Spreadsheet with dependencies saved to {output_xlsx}")

if __name__ == "__main__":
    # Input IR file
    ir_file = "../test/sample.ll"  # Specify the path to your IR file
    
    # Extract functions from the IR
    functions = extract_functions_from_ir(ir_file)

    # Output Excel file
    output_xlsx = "../data/processed/ir_instructions_with_cell_deps.xlsx"

    # Convert functions to a spreadsheet with cell-based dependencies
    convert_to_spreadsheet_with_dependency(functions, output_xlsx)
